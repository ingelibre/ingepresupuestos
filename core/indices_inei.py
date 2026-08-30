# SPDX-License-Identifier: GPL-3.0-or-later
# Copyright (C) 2026 Marco Sumari
# This file is part of IngePresupuestos — https://ingepresupuestos.com
# Software libre bajo la GNU GPL v3 o posterior. Ver el archivo LICENSE.
"""core.indices_inei — gestión del histórico de Índices Unificados de Precios (INEI).

Tablas:
    - ``indices_inei``         : catálogo de los 80 códigos
    - ``indices_inei_areas``   : 6 áreas geográficas estándar
    - ``indices_inei_valores`` : serie histórica (codigo · año · mes · área · valor)

Soporta importación masiva desde Excel publicado por el INEI cada mes y
exportación / importación de la serie completa en JSON para sincronización
entre instalaciones.

Espejo conceptual de la funcionalidad "Importación de Índices de Precios
INEI 2026" de Delphin Express.
"""
from __future__ import annotations

import json
from pathlib import Path

from core.database import get_db


# ── Series de índices unificados ─────────────────────────────────────────────
# El INEI cambió la base con la RJ 016-2026-INEI (20-01-2026): de «Julio 1992 =
# 100» con 6 áreas geográficas a «Diciembre 2025 = 100» con 13. No es solo un
# cambio de escala: 30 de los códigos comunes CAMBIARON DE SIGNIFICADO —el 21
# era «Cemento Portland Tipo I» y ahora es «Cemento Portland e hidráulico»,
# que absorbió al 22 y al 23, desaparecidos— y se agregaron 15 (81 a 95).
#
# Por eso conviven las dos series y NO se mezclan: leer una fórmula de 2024 con
# la tabla nueva daría un número equivocado sin avisar.
SERIE_2025 = '2025'
SERIE_1992 = '1992'
SERIE_ACTUAL = SERIE_2025

# Primer período de la serie nueva. Antes de esto manda la de 1992.
INICIO_SERIE_2025 = (2025, 12)

_OFICIAL: dict | None = None


def _oficial() -> dict:
    """Los anexos de la RJ 016-2026-INEI, leídos del recurso empaquetado.

    `resources/indices_inei_oficial.json` trae las dos relaciones de índices,
    las áreas de cada serie y el Diccionario de Elementos de la Construcción.
    Se carga una vez y se cachea.
    """
    global _OFICIAL
    if _OFICIAL is None:
        try:
            from core.config import BASE_DIR
            ruta = Path(BASE_DIR) / "resources" / "indices_inei_oficial.json"
            with open(ruta, encoding="utf-8") as f:
                _OFICIAL = json.load(f)
        except Exception:
            # Sin el recurso la app sigue viva con la serie histórica.
            _OFICIAL = {'series': {}, 'diccionario': {}}
    return _OFICIAL


_VALORES_OFICIALES = None


def _valores_oficiales() -> dict:
    """El histórico de valores empaquetado, `indices_inei_valores.json.gz`.

    Lo genera `scripts/generar_indices_valores.py` antes de cada release desde
    las fuentes del INEI. Existe porque los índices cambian todos los meses y
    un release sale cada tanto: sin él, **una instalación nueva arranca sin la
    base vigente** —el Excel del INEI lleva meses congelado y gob.pe solo deja
    el PDF del mes en curso— y no hay reajuste que calcular.

    Formato compacto: por serie, `areas` y `datos[período][código] = [v, …]`
    con una posición por área (`null` donde el índice no existe). Guardar un
    dict por valor multiplicaba por diez el tamaño; así los 66 000 valores
    ocupan 72 KB.
    """
    global _VALORES_OFICIALES
    if _VALORES_OFICIALES is None:
        try:
            import gzip
            from core.config import BASE_DIR
            ruta = (Path(BASE_DIR) / "resources"
                    / "indices_inei_valores.json.gz")
            with gzip.open(ruta, 'rt', encoding='utf-8') as f:
                _VALORES_OFICIALES = json.load(f)
        except Exception:
            # Sin el recurso la app sigue viva: se sincroniza y ya.
            _VALORES_OFICIALES = {'series': {}}
    return _VALORES_OFICIALES


def series_disponibles() -> list[tuple[str, str]]:
    """[(clave, nombre)] de las series, la más reciente primero."""
    s = _oficial().get('series') or {}
    out = [(k, v.get('nombre', k)) for k, v in s.items()]
    if not out:
        out = [(SERIE_1992, 'Base Julio 1992 = 100')]
    return sorted(out, key=lambda x: x[0], reverse=True)


def serie_de(anio, mes) -> str:
    """Qué serie corresponde a un período. Es la regla que evita mezclarlas."""
    try:
        return (SERIE_2025 if (int(anio), int(mes)) >= INICIO_SERIE_2025
                else SERIE_1992)
    except (TypeError, ValueError):
        return SERIE_ACTUAL


def diccionario_oficial() -> dict[str, str]:
    """Anexo 2: elemento de construcción → código de índice unificado.

    Son ~1930 entradas publicadas por el INEI, la referencia con autoridad para
    decidir qué índice le toca a un insumo.
    """
    return _oficial().get('diccionario') or {}


# ── Semilla del catálogo de índices unificados ───────────────────────────────
# Respaldo de la serie 1992 por si falta el recurso: 72 entradas con códigos del
# 01 al 80 (la numeración tiene huecos). La verdad, igual que antes, es la tabla
# `indices_inei`; para leerla usar `catalogo()`, NUNCA esta constante.
CATALOGO_INEI: list[tuple[str, str]] = [
    ("01", "Aceite"),
    ("02", "Acero de construcción liso"),
    ("03", "Acero de construcción corrugado"),
    ("04", "Agregado fino"),
    ("05", "Agregado grueso"),
    ("06", "Alambre y cable de cobre desnudo"),
    ("07", "Alambre y cable tipo TW y THW"),
    ("08", "Alambre y cable tipo WP"),
    ("09", "Alcantarilla metálica"),
    ("10", "Aparato sanitario con grifería"),
    ("11", "Artefacto de alumbrado exterior"),
    ("12", "Artefacto de alumbrado interior"),
    ("13", "Asfalto"),
    ("14", "Baldosa acústica"),
    ("15", "Baldosa asfáltica"),
    ("16", "Baldosa vinílica"),
    ("17", "Bloque y ladrillo"),
    ("18", "Cable telefónico"),
    ("19", "Cable NYY-N2XY"),
    ("20", "Cemento asfáltico"),
    ("21", "Cemento Portland tipo I"),
    ("22", "Cemento Portland tipo II"),
    ("23", "Cemento Portland tipo V"),
    ("24", "Cerámica esmaltada y sin esmaltar"),
    ("26", "Cerrajería nacional"),
    ("27", "Detonante"),
    ("28", "Dinamita"),
    ("29", "Dólar"),
    ("30", "Dólar más inflación USA / General ponderado"),
    ("31", "Ducto de concreto"),
    ("32", "Flete terrestre"),
    ("33", "Flete aéreo"),
    ("34", "Gasolina"),
    ("37", "Herramienta manual"),
    ("38", "Hormigón"),
    ("39", "Índice general de precios al consumidor (IPC)"),
    ("40", "Loseta"),
    ("41", "Madera en tiras para piso"),
    ("42", "Madera importada para encofrado y carpintería"),
    ("43", "Madera nacional para encofrado y carpintería"),
    ("44", "Madera terciada para encofrado y carpintería"),
    ("45", "Madera terciada para encofrado"),
    ("46", "Malla de acero"),
    ("47", "Mano de obra (incluido leyes sociales)"),
    ("48", "Maquinaria y equipo nacional"),
    ("49", "Maquinaria y equipo importado"),
    ("50", "Marco y tapa de hierro fundido"),
    ("51", "Perfil de acero liviano"),
    ("52", "Perfil de aluminio"),
    ("53", "Petróleo diesel"),
    ("54", "Pintura látex"),
    ("55", "Pintura temple"),
    ("56", "Plancha de Aero LAC"),
    ("57", "Plancha de Aero LAF"),
    ("59", "Plancha de fibro-cemento"),
    ("60", "Plancha de poliuretano"),
    ("61", "Plancha galvanizada"),
    ("62", "Poste de concreto"),
    ("64", "Terrazo"),
    ("65", "Tubería de acero negro y/o galvanizado"),
    ("66", "Tubería de PVC para agua potable y alcantarillado"),
    ("68", "Tubería de cobre"),
    ("69", "Tubería de concreto simple"),
    ("70", "Tubería de concreto reforzado"),
    ("71", "Tubería de fierro fundido"),
    ("72", "Tubería de PVC para agua"),
    ("73", "Ducto telefónico de PVC"),
    ("74", "Tubería de PVC para electricidad (SAP)"),
    ("77", "Válvula de bronce nacional"),
    ("78", "Válvula de fierro fundido nacional"),
    ("79", "Vidrio incoloro nacional"),
    ("80", "Concreto premezclado"),
]
AREAS_INEI: list[tuple[str, str]] = [
    ("01", "Lima Metropolitana y Callao"),
    ("02", "Norte (Tumbes/Piura/Lambayeque/La Libertad/Áncash)"),
    ("03", "Centro (Lima Provincias/Junín/Pasco/Huánuco)"),
    ("04", "Sur Medio y Selva (Ica/Ayacucho/Huancavelica/Amazonas/San Martín/Loreto)"),
    ("05", "Sur (Arequipa/Moquegua/Tacna/Apurímac/Cusco/Madre de Dios/Puno)"),
    ("06", "Nacional (promedio ponderado)"),
]


# ── Semilla del catálogo ─────────────────────────────────────────────────────
# La semilla corre UNA vez por serie y por versión. Subir este número al
# corregir o ampliar los datos oficiales: solo entonces se vuelve a sembrar.
SEED_VERSION = 2
# Sube al regenerar `indices_inei_valores.json.gz` con más meses: es lo
# que hace que una instalación ya existente reciba lo nuevo.
VALORES_VERSION = 2


def _seed_de(serie: str) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    """(índices, áreas) oficiales de una serie. Cae al respaldo de 1992."""
    s = (_oficial().get('series') or {}).get(serie)
    if s:
        indices = sorted((s.get('indices') or {}).items(), key=lambda kv: kv[0])
        areas = [tuple(a) for a in (s.get('areas') or [])]
        if indices:
            return indices, areas
    if serie == SERIE_1992:
        return list(CATALOGO_INEI), list(AREAS_INEI)
    return [], []


def asegurar_seed(conn=None, serie: str | None = None) -> None:
    """Siembra el catálogo y las áreas de cada serie. Una vez por serie.

    Antes re-insertaba las 72 entradas en CADA arranque con INSERT OR IGNORE,
    así que un índice borrado por el usuario resucitaba al reiniciar. Ahora la
    siembra se salta si `seed_inei_<serie>` ya alcanzó SEED_VERSION, y los
    borrados y renombres del usuario mandan.

    El flag es POR SERIE: incorporar la base 2025 no debe deshacer lo que el
    usuario haya tocado en la de 1992.

    OJO: el alta, la edición y la baja la llaman ANTES de escribir. Si no, un
    borrado hecho sobre una BD donde la semilla todavía no corrió se deshacía
    solo en la siguiente lectura.
    """
    own = conn is None
    if own:
        conn = get_db()
    try:
        series = [serie] if serie else [SERIE_1992, SERIE_2025]
        for s in series:
            clave = f"seed_inei_{s}"
            row = conn.execute(
                "SELECT valor FROM configuracion WHERE clave=?", (clave,)
            ).fetchone()
            ya = int((row['valor'] if row else 0) or 0)
            if ya >= SEED_VERSION:
                continue
            indices, areas = _seed_de(s)
            for codigo, nombre in indices:
                conn.execute(
                    "INSERT OR IGNORE INTO indices_inei (codigo, serie, nombre, activo)"
                    " VALUES (?,?,?,1)", (codigo, s, nombre)
                )
            for i, (codigo, nombre) in enumerate(areas):
                conn.execute(
                    "INSERT OR IGNORE INTO indices_inei_areas "
                    "(codigo, serie, nombre, orden) VALUES (?,?,?,?)",
                    (codigo, s, nombre, i)
                )
            conn.execute(
                "INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?,?)",
                (clave, str(SEED_VERSION))
            )
        for s_ in series:
            _sembrar_valores(conn, s_)
        conn.commit()
        if not serie:                 # solo en el arranque, no en cada alta
            refrescar_valores_oficiales(conn)
    finally:
        if own:
            conn.close()


def _sembrar_valores(conn, serie: str) -> int:
    """Vuelca el histórico empaquetado de una serie. Una vez por versión.

    **`INSERT OR IGNORE`, nunca REPLACE**: si el usuario ya sincronizó o
    corrigió un valor a mano, el suyo manda. Esto solo rellena lo que falta.

    Con su propio flag, aparte del catálogo: una instalación que ya venía
    funcionando tiene `seed_inei_<serie>` al día y aun así necesita recibir los
    valores la primera vez que actualiza a una versión que los trae.
    """
    clave = f"seed_inei_valores_{serie}"
    row = conn.execute("SELECT valor FROM configuracion WHERE clave=?",
                       (clave,)).fetchone()
    if int((row['valor'] if row else 0) or 0) >= VALORES_VERSION:
        return 0
    bloque = (_valores_oficiales().get('series') or {}).get(serie) or {}
    n = 0
    for f in _filas_de_bloque(serie, bloque):
        conn.execute(
            "INSERT OR IGNORE INTO indices_inei_valores "
            "(codigo, serie, anio, mes, area, valor) VALUES (?,?,?,?,?,?)",
            (f['codigo'], serie, f['anio'], f['mes'], f['area'], f['valor'])
        )
        n += 1
    conn.execute(
        "INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?,?)",
        (clave, str(VALORES_VERSION))
    )
    return n


def refrescar_valores_oficiales(conn=None) -> int:
    """Corrige de una sola vez los índices que el seed viejo dejó mal.

    Distinto de `_sembrar_valores`, que rellena sin pisar. Esto **pisa**, y
    corre UNA vez en la vida de cada base: hasta la 3.0.4 el seed traía 2 212
    valores que contradecían al INEI —entre ellos marcadores como 100.00,
    500.00 y 1000.00 en 2024— y como la siembra ignora lo que ya existe, esa
    basura se quedaba para siempre en toda instalación que ya venía
    funcionando. Un reajuste calculado con esos números sale mal y nadie se
    entera.

    Solo toca valores que el archivo oficial del INEI publica y que difieren
    de lo guardado. Después de esta pasada vuelve a mandar el usuario: la
    siembra normal nunca pisa nada, y una corrección a mano queda.
    """
    own = conn is None
    if own:
        conn = get_db()
    try:
        clave = "indices_refresco_oficial"
        row = conn.execute("SELECT valor FROM configuracion WHERE clave=?",
                           (clave,)).fetchone()
        if row and str(row['valor'] or '') == '1':
            return 0
        n = 0
        for serie, bloque in (_valores_oficiales().get('series') or {}).items():
            areas = bloque.get('areas') or []
            for periodo, codigos in (bloque.get('datos') or {}).items():
                try:
                    anio, mes = (int(x) for x in periodo.split('-'))
                except ValueError:
                    continue
                for codigo, valores in codigos.items():
                    for i, valor in enumerate(valores):
                        if valor is None or i >= len(areas):
                            continue
                        cur = conn.execute(
                            "SELECT valor FROM indices_inei_valores WHERE "
                            "codigo=? AND serie=? AND anio=? AND mes=? AND area=?",
                            (codigo, serie, anio, mes, areas[i])).fetchone()
                        if cur is None or abs(cur['valor'] - float(valor)) < 0.005:
                            continue
                        conn.execute(
                            "UPDATE indices_inei_valores SET valor=? WHERE "
                            "codigo=? AND serie=? AND anio=? AND mes=? AND area=?",
                            (float(valor), codigo, serie, anio, mes, areas[i]))
                        n += 1
        # La base Julio 1992 dejó de existir en diciembre de 2025: cualquier
        # valor suyo posterior es residuo de un import viejo y confunde al
        # mirar el histórico.
        n += conn.execute(
            "DELETE FROM indices_inei_valores WHERE serie=? AND "
            "(anio*12+mes) > ?", (SERIE_1992,
                                  INICIO_SERIE_2025[0] * 12 + INICIO_SERIE_2025[1])
        ).rowcount or 0
        conn.execute(
            "INSERT OR REPLACE INTO configuracion (clave, valor) VALUES (?,'1')",
            (clave,))
        conn.commit()
        return n
    finally:
        if own:
            conn.close()


# ── Catálogo: alta, edición y baja ───────────────────────────────────────────
def _norm_codigo(codigo) -> str:
    """Normaliza a la forma canónica del INEI: dos dígitos, '7' → '07'."""
    s = str(codigo or '').strip()
    if not s.isdigit():
        raise ValueError("El código debe ser numérico (01 a 99).")
    n = int(s)
    if not (0 <= n <= 99):
        raise ValueError("El código debe estar entre 00 y 99.")
    return f"{n:02d}"


def catalogo(incluir_inactivos: bool = False, conn=None,
             serie: str = SERIE_ACTUAL) -> list[tuple[str, str]]:
    """El catálogo vigente de una serie, leído de la tabla.

    Fuente única para toda la app: las vistas lo usan para poblar sus combos,
    así que un índice dado de alta aparece en todas sin tocar código.
    """
    own = conn is None
    if own:
        conn = get_db()
    asegurar_seed(conn)
    sql = "SELECT codigo, nombre FROM indices_inei WHERE serie=?"
    if not incluir_inactivos:
        sql += " AND activo=1"
    sql += " ORDER BY codigo"
    rows = conn.execute(sql, (serie,)).fetchall()
    if own:
        conn.close()
    return [(r['codigo'], r['nombre']) for r in rows]


def crear_indice(codigo: str, nombre: str, conn=None,
                 serie: str = SERIE_ACTUAL) -> str:
    """Da de alta un índice unificado. Devuelve el código normalizado."""
    codigo = _norm_codigo(codigo)
    nombre = str(nombre or '').strip()
    if not nombre:
        raise ValueError("El nombre no puede estar vacío.")
    own = conn is None
    if own:
        conn = get_db()
    try:
        asegurar_seed(conn)
        ya = conn.execute(
            "SELECT nombre FROM indices_inei WHERE codigo=? AND serie=?",
            (codigo, serie)
        ).fetchone()
        if ya:
            raise ValueError(f"El índice {codigo} ya existe ({ya['nombre']}).")
        conn.execute(
            "INSERT INTO indices_inei (codigo, serie, nombre, activo)"
            " VALUES (?,?,?,1)", (codigo, serie, nombre)
        )
        conn.commit()
    finally:
        if own:
            conn.close()
    return codigo


def actualizar_indice(codigo: str, nombre: str | None = None,
                      activo: bool | None = None, conn=None,
                      serie: str = SERIE_ACTUAL) -> None:
    """Renombra o activa/desactiva un índice ya existente."""
    codigo = _norm_codigo(codigo)
    sets, params = [], []
    if nombre is not None:
        nombre = str(nombre).strip()
        if not nombre:
            raise ValueError("El nombre no puede estar vacío.")
        sets.append("nombre=?")
        params.append(nombre)
    if activo is not None:
        sets.append("activo=?")
        params.append(1 if activo else 0)
    if not sets:
        return
    params.extend([codigo, serie])
    own = conn is None
    if own:
        conn = get_db()
    try:
        asegurar_seed(conn)
        conn.execute(
            f"UPDATE indices_inei SET {', '.join(sets)} "
            f"WHERE codigo=? AND serie=?", params
        )
        conn.commit()
    finally:
        if own:
            conn.close()


def contar_usos(codigo: str, conn=None, serie: str = SERIE_ACTUAL) -> dict:
    """Qué quedaría colgando si se borra este índice.

    No hay clave foránea desde `recursos` ni desde `formula_monomios`, así que
    borrar no rompe nada a nivel SQL — pero deja insumos apuntando a un código
    que ya no existe. La vista usa esto para avisar antes.
    """
    codigo = _norm_codigo(codigo)
    own = conn is None
    if own:
        conn = get_db()
    try:
        recursos = conn.execute(
            "SELECT COUNT(*) FROM recursos WHERE indice_inei=?", (codigo,)
        ).fetchone()[0]
        valores = conn.execute(
            "SELECT COUNT(*) FROM indices_inei_valores WHERE codigo=? AND serie=?",
            (codigo, serie)
        ).fetchone()[0]
        monomios = conn.execute(
            "SELECT COUNT(*) FROM formula_monomios WHERE indice_inei=?", (codigo,)
        ).fetchone()[0]
    finally:
        if own:
            conn.close()
    return {'recursos': recursos, 'valores': valores, 'monomios': monomios}


def eliminar_indice(codigo: str, borrar_valores: bool = False, conn=None,
                    serie: str = SERIE_ACTUAL) -> None:
    """Baja del catálogo. Con `borrar_valores`, se lleva también su histórico.

    Los insumos que lo usaban NO se tocan: conservan el código para que el
    usuario decida a dónde reasignarlos.
    """
    codigo = _norm_codigo(codigo)
    own = conn is None
    if own:
        conn = get_db()
    try:
        asegurar_seed(conn)
        conn.execute("DELETE FROM indices_inei WHERE codigo=? AND serie=?",
                     (codigo, serie))
        if borrar_valores:
            conn.execute(
                "DELETE FROM indices_inei_valores WHERE codigo=? AND serie=?",
                (codigo, serie)
            )
        conn.commit()
    finally:
        if own:
            conn.close()


def asegurar_codigos(codigos, nombres: dict | None = None, conn=None,
                     serie: str = SERIE_ACTUAL) -> int:
    """Da de alta los códigos que aún no estén en el catálogo. Devuelve cuántos.

    Es lo que hace utilizable la importación del archivo oficial: hasta ahora
    los valores de un código ausente del catálogo SÍ se guardaban en
    `indices_inei_valores` —esa tabla no tiene clave foránea— pero la lista se
    arma desde `indices_inei`, así que quedaban invisibles e inservibles.
    """
    nombres = nombres or {}
    own = conn is None
    if own:
        conn = get_db()
    nuevos = 0
    try:
        for c in codigos:
            try:
                cod = _norm_codigo(c)
            except ValueError:
                continue
            ya = conn.execute(
                "SELECT 1 FROM indices_inei WHERE codigo=? AND serie=?",
                (cod, serie)
            ).fetchone()
            if ya:
                continue
            nombre = str(nombres.get(cod) or '').strip() or f"Índice {cod}"
            conn.execute(
                "INSERT INTO indices_inei (codigo, serie, nombre, activo)"
                " VALUES (?,?,?,1)", (cod, serie, nombre)
            )
            nuevos += 1
        conn.commit()
    finally:
        if own:
            conn.close()
    return nuevos


def codigos_huerfanos(conn=None, serie: str = SERIE_ACTUAL) -> list[dict]:
    """Códigos que la app usa pero el catálogo de la serie no define.

    Los clasifica, porque la respuesta correcta es distinta en cada caso:

    * ``descontinuado`` — el código SÍ existía en la relación anterior y el
      INEI lo retiró al reagruparlo (el 22 y el 23, «Cemento Portland Tipo II»
      y «Tipo V», los absorbió el 21). Darlo de alta en la serie nueva sería
      inventar un índice que el INEI no publica y que nunca tendrá valores: lo
      que corresponde es REASIGNAR esos insumos a un código vigente.
    * ``desconocido`` — no figura en ninguna relación oficial. Suele venir de
      bibliotecas importadas: el 99, por ejemplo, son subcontratos.

    El '00' se excluye: no es un índice del INEI sino el centinela que usa
    `core.parte_diario` para lo sin clasificar.
    """
    own = conn is None
    if own:
        conn = get_db()
    try:
        rows = conn.execute(
            """SELECT codigo, SUM(n_recursos) AS n_recursos, SUM(n_valores) AS n_valores
                 FROM (
                   SELECT indice_inei AS codigo, COUNT(*) AS n_recursos, 0 AS n_valores
                     FROM recursos
                    WHERE COALESCE(indice_inei,'') NOT IN ('', '00')
                      AND indice_inei NOT IN (
                          SELECT codigo FROM indices_inei WHERE serie=?)
                    GROUP BY indice_inei
                   UNION ALL
                   SELECT codigo, 0, COUNT(*)
                     FROM indices_inei_valores
                    WHERE serie=? AND codigo NOT IN ('00')
                      AND codigo NOT IN (
                          SELECT codigo FROM indices_inei WHERE serie=?)
                    GROUP BY codigo
                 )
                GROUP BY codigo ORDER BY codigo""", (serie, serie, serie)
        ).fetchall()
    finally:
        if own:
            conn.close()

    # Relaciones OFICIALES de las otras series, para saber si el código fue
    # descontinuado o nunca existió.
    otras = {}
    for clave, datos in (_oficial().get('series') or {}).items():
        if clave == serie:
            continue
        for cod, nombre in (datos.get('indices') or {}).items():
            otras.setdefault(cod, (clave, nombre))

    out = []
    for r in rows:
        d = dict(r)
        anterior = otras.get(d['codigo'])
        d['descontinuado'] = anterior is not None
        d['serie_anterior'] = anterior[0] if anterior else ''
        d['nombre_anterior'] = anterior[1] if anterior else ''
        out.append(d)
    return out


# ── Listados ─────────────────────────────────────────────────────────────────
def listar_indices(conn=None, serie: str = SERIE_ACTUAL) -> list[dict]:
    """Devuelve el catálogo con el último valor cargado (de cualquier área)."""
    own = conn is None
    if own:
        conn = get_db()
    asegurar_seed(conn)
    rows = conn.execute(
        """SELECT i.codigo, i.nombre, i.activo,
                  (SELECT COUNT(DISTINCT anio || '-' || mes || '-' || area)
                   FROM indices_inei_valores v
                   WHERE v.codigo = i.codigo AND v.serie = i.serie)
                  AS n_valores,
                  (SELECT anio || '-' || PRINTF('%02d', mes)
                   FROM indices_inei_valores v
                   WHERE v.codigo = i.codigo AND v.serie = i.serie
                   ORDER BY anio DESC, mes DESC LIMIT 1)
                  AS ultimo_periodo,
                  (SELECT valor FROM indices_inei_valores v
                   WHERE v.codigo = i.codigo AND v.serie = i.serie
                   ORDER BY anio DESC, mes DESC LIMIT 1)
                  AS ultimo_valor
           FROM indices_inei i WHERE i.serie=? ORDER BY i.codigo""", (serie,)
    ).fetchall()
    if own:
        conn.close()
    return [dict(r) for r in rows]


def listar_areas(conn=None, serie: str = SERIE_ACTUAL) -> list[dict]:
    """Las áreas geográficas de la serie: 6 en la de 1992, 13 en la de 2025."""
    own = conn is None
    if own:
        conn = get_db()
    asegurar_seed(conn)
    rows = conn.execute(
        "SELECT codigo, nombre, orden FROM indices_inei_areas "
        "WHERE serie=? ORDER BY orden", (serie,)
    ).fetchall()
    if own:
        conn.close()
    return [dict(r) for r in rows]


def obtener_valor(codigo: str, anio: int, mes: int,
                  area: str = '01', serie: str | None = None) -> float | None:
    """Valor de un índice en un período. La serie se deduce de la fecha.

    Deducirla acá es lo que impide el error silencioso de leer un período de
    2024 contra la base Diciembre 2025.
    """
    serie = serie or serie_de(anio, mes)
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT valor FROM indices_inei_valores "
            "WHERE codigo=? AND serie=? AND anio=? AND mes=? AND area=?",
            (str(codigo), serie, int(anio), int(mes), str(area))
        ).fetchone()
    finally:
        conn.close()
    return row['valor'] if row else None


def obtener_matriz(codigo: str, area: str = '01',
                   serie: str = SERIE_ACTUAL) -> dict[int, dict[int, float]]:
    """Histórico de un índice como {año: {mes: valor}}."""
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT anio, mes, valor FROM indices_inei_valores "
            "WHERE codigo=? AND serie=? AND area=? ORDER BY anio, mes",
            (str(codigo), serie, str(area))
        ).fetchall()
    finally:
        conn.close()
    out: dict[int, dict[int, float]] = {}
    for r in rows:
        out.setdefault(r['anio'], {})[r['mes']] = r['valor']
    return out


# ── Persistencia ─────────────────────────────────────────────────────────────
def guardar_valor(codigo: str, anio: int, mes: int, valor: float,
                  area: str = '01', serie: str | None = None) -> None:
    """Inserta o reemplaza un valor. La serie sale de la fecha si no se indica."""
    serie = serie or serie_de(anio, mes)
    conn = get_db()
    try:
        conn.execute(
            "INSERT OR REPLACE INTO indices_inei_valores "
            "(codigo, serie, anio, mes, area, valor) VALUES (?,?,?,?,?,?)",
            (codigo, serie, int(anio), int(mes), area, float(valor))
        )
        conn.commit()
    finally:
        conn.close()


def guardar_valores(rows: list[dict]) -> tuple[int, int]:
    """Batch upsert. ``rows`` lleva codigo/anio/mes/area/valor y opcional serie.
    Retorna (n_insertados_o_actualizados, n_ignorados_por_error)."""
    conn = get_db()
    ok = 0
    err = 0
    try:
        # Alta automática de los códigos que el catálogo aún no tenga: esta
        # tabla no tiene clave foránea, así que sin esto los valores entraban
        # pero el índice quedaba invisible en la lista.
        por_serie: dict[str, set] = {}
        nombres_por_serie: dict[str, dict] = {}
        for r in rows:
            cod = str(r.get('codigo') or '').strip().zfill(2)
            s = r.get('serie') or serie_de(r.get('anio'), r.get('mes'))
            por_serie.setdefault(s, set()).add(cod)
            if r.get('nombre'):
                nombres_por_serie.setdefault(s, {})[cod] = r['nombre']
        for s, cods in por_serie.items():
            asegurar_codigos(cods, nombres_por_serie.get(s, {}), conn, serie=s)

        for r in rows:
            try:
                codigo = str(r.get('codigo') or '').strip().zfill(2)[:2]
                anio = int(r.get('anio') or 0)
                mes = int(r.get('mes') or 0)
                area = str(r.get('area') or '01')
                valor = float(r.get('valor') or 0)
                serie = r.get('serie') or serie_de(anio, mes)
                if not codigo or anio < 1900 or not (1 <= mes <= 12) or valor <= 0:
                    err += 1
                    continue
                conn.execute(
                    "INSERT OR REPLACE INTO indices_inei_valores "
                    "(codigo, serie, anio, mes, area, valor) VALUES (?,?,?,?,?,?)",
                    (codigo, serie, anio, mes, area, valor)
                )
                ok += 1
            except Exception:
                err += 1
        conn.commit()
    finally:
        conn.close()
    return ok, err


def eliminar_valor(codigo: str, anio: int, mes: int, area: str = '01',
                   serie: str | None = None) -> None:
    serie = serie or serie_de(anio, mes)
    conn = get_db()
    try:
        conn.execute(
            "DELETE FROM indices_inei_valores "
            "WHERE codigo=? AND serie=? AND anio=? AND mes=? AND area=?",
            (codigo, serie, anio, mes, area)
        )
        conn.commit()
    finally:
        conn.close()


# ── Importación / Exportación ────────────────────────────────────────────────
MESES_MAP = {
    'enero': 1, 'ene': 1, 'jan': 1, 'january': 1,
    'febrero': 2, 'feb': 2, 'february': 2,
    'marzo': 3, 'mar': 3, 'march': 3,
    'abril': 4, 'abr': 4, 'apr': 4, 'april': 4,
    'mayo': 5, 'may': 5,
    'junio': 6, 'jun': 6, 'june': 6,
    'julio': 7, 'jul': 7, 'july': 7,
    'agosto': 8, 'ago': 8, 'aug': 8, 'august': 8,
    # 'set' es la abreviatura que usa el INEI en sus hojas («Set-2013»),
    # y sin ella se perdía SEPTIEMBRE de todos los años del acumulativo.
    'septiembre': 9, 'setiembre': 9, 'sep': 9, 'sept': 9, 'set': 9,
    'september': 9,
    'octubre': 10, 'oct': 10, 'october': 10,
    'noviembre': 11, 'nov': 11, 'november': 11,
    'diciembre': 12, 'dic': 12, 'dec': 12, 'december': 12,
}


def _parse_mes(texto) -> int | None:
    """Resuelve un encabezado de columna a número de mes (1-12).

    Acepta números (1, 01, 1.0), nombres ('Ene', 'Enero', 'JAN') con o sin
    tildes y acentos. Retorna None si no se puede resolver.
    """
    if texto is None:
        return None
    s = str(texto).strip().lower()
    if not s:
        return None
    s = (s.replace('á', 'a').replace('é', 'e').replace('í', 'i')
           .replace('ó', 'o').replace('ú', 'u'))
    if s in MESES_MAP:
        return MESES_MAP[s]
    try:
        v = int(float(s))
        if 1 <= v <= 12:
            return v
    except Exception:
        pass
    return None


def _parse_hoja_periodo(nombre: str):
    """(anio, mes) del nombre de una hoja del archivo INEI, o None.

    Las hojas se llaman «Ene_2026», «Abr_2021», «Dic-2025»… y conviven con
    hojas que NO son meses («Relación de Indices», «Diccionario Alfabetico»,
    «Mano_de_obra_2026»), que hay que saltar.
    """
    import re as _re
    s = str(nombre or '').strip()
    m = _re.match(r'^\s*([A-Za-zÁÉÍÓÚáéíóú]{3,12})[\s_\-.]+(\d{4})\s*$', s)
    if not m:
        return None
    mes = _parse_mes(m.group(1))
    if not mes:
        return None
    return int(m.group(2)), mes


def _bloques_de_areas(fila) -> list[tuple[int, dict[int, str]]]:
    """Localiza los bloques «Cód. | 1 | 2 | …» de una fila de encabezado.

    El archivo del INEI pone DOS tablas lado a lado —códigos impares a la
    izquierda, pares a la derecha— y cada una lleva sus columnas de áreas
    numeradas 1..6 (base 1992) o 1..13 (base 2025).

    Devuelve [(col_del_codigo, {col_valor: codigo_de_area})].
    """
    bloques = []
    actual = None
    for j, val in enumerate(fila):
        if val is None:
            continue
        s = str(val).strip().lower().rstrip('.')
        if s.startswith('cod') or s.startswith('cód'):
            actual = (j, {})
            bloques.append(actual)
            continue
        if actual is None:
            continue
        try:
            n = int(float(str(val).strip()))
        except (TypeError, ValueError):
            continue
        if 1 <= n <= 20:
            actual[1][j] = f"{n:02d}"
    return [b for b in bloques if b[1]]


def _serie_del_libro(wb) -> str | None:
    """Lee la base declarada en la cabecera: «(Base : Diciembre 2025 = 100)»."""
    for ws in list(wb.worksheets)[:3]:
        for row in ws.iter_rows(values_only=True, max_row=8):
            for v in row:
                if v is None:
                    continue
                t = str(v).lower()
                if 'base' in t:
                    if '2025' in t:
                        return SERIE_2025
                    if '1992' in t:
                        return SERIE_1992
    return None


def _importar_oficial(wb, serie_forzada: str | None = None) -> dict | None:
    """Lee el archivo del INEI con su estructura real. None si no es ese formato.

    Estructura: UNA HOJA POR MES, y dentro cada columna es un ÁREA GEOGRÁFICA.
    El lector anterior hacía justo lo contrario —tomaba `wb.active` y trataba
    los números 1..6 del encabezado como MESES—, así que de un archivo de 111
    hojas sacaba una sola, mezclaba las áreas de dos índices distintos como si
    fueran los 12 meses de uno, y lo guardaba todo bajo el área que el usuario
    hubiera elegido en el combo. Los valores con los que se calculaba K eran
    inventados.
    """
    serie = serie_forzada or _serie_del_libro(wb) or SERIE_ACTUAL
    rows_out: list[dict] = []
    ignorados = 0
    codigos: set[str] = set()
    hojas = 0

    for ws in wb.worksheets:
        per = _parse_hoja_periodo(ws.title)
        if not per:
            continue
        anio, mes = per
        datos = list(ws.iter_rows(values_only=True))
        bloques = None
        for fila in datos[:20]:
            b = _bloques_de_areas(fila)
            if b and sum(len(x[1]) for x in b) >= 4:
                bloques = b
                break
        if not bloques:
            continue
        hojas += 1
        for fila in datos:
            for col_cod, cols_area in bloques:
                if col_cod >= len(fila):
                    continue
                cod_raw = fila[col_cod]
                if cod_raw is None:
                    continue
                s = str(cod_raw).strip().replace('.0', '')
                if not s.isdigit() or not (1 <= int(s) <= 99):
                    continue
                codigo = s.zfill(2)
                for col, area in cols_area.items():
                    if col >= len(fila):
                        continue
                    v = fila[col]
                    if v is None or str(v).strip() in ('', '(*)', '*'):
                        continue          # «(*) Sin índice» — no es un cero
                    try:
                        f = float(v)
                    except (TypeError, ValueError):
                        ignorados += 1
                        continue
                    if f <= 0:
                        ignorados += 1
                        continue
                    rows_out.append({'codigo': codigo, 'serie': serie,
                                     'anio': anio, 'mes': mes,
                                     'area': area, 'valor': f})
                    codigos.add(codigo)

    if not rows_out:
        return None
    periodos = {(r['anio'], r['mes']) for r in rows_out}
    areas = {r['area'] for r in rows_out}
    return {
        'ok': True,
        'formato': 'oficial',
        'serie': serie,
        'msg': (f"Archivo oficial del INEI ({serie_nombre(serie)}): "
                f"{len(rows_out)} valores · {len(codigos)} índices · "
                f"{len(periodos)} meses · {len(areas)} áreas."),
        'rows': rows_out,
        'ignorados': ignorados,
        'anio_detectado': max(a for a, _ in periodos),
        'codigos_encontrados': codigos,
        'nombres_encontrados': _relacion_del_libro(wb),
        'hojas': hojas,
        'periodos': len(periodos),
        'areas': sorted(areas),
    }


def _relacion_del_libro(wb) -> dict[str, str]:
    """La hoja «Relación de Índices» del propio archivo, si viene."""
    import re as _re
    for ws in wb.worksheets:
        if 'relaci' not in ws.title.lower():
            continue
        out = {}
        for row in ws.iter_rows(values_only=True):
            c = list(row) + [None] * 8
            for a, b in ((0, 1), (1, 2), (2, 3), (4, 5)):
                cod, nom = c[a], c[b]
                if cod is None or nom is None:
                    continue
                s = str(cod).strip().replace('.0', '')
                if _re.fullmatch(r'\d{1,3}', s):
                    n = ' '.join(str(nom).split())
                    if n and not _re.fullmatch(r'[\d.,]+', n):
                        out[s.zfill(2)] = _re.sub(r'\s*\([a-z]\)\s*$', '', n)
        if out:
            return out
    return {}


def serie_nombre(serie: str) -> str:
    return dict(series_disponibles()).get(serie, serie)


def _nombre_en_fila(row, cod_str: str, codigo_col: int, mes_cols: dict) -> str:
    """Saca la descripción del índice de una fila con formato libre.

    Dos formas, en orden: pegada al código en la misma celda ('85 - CABLE…')
    o en alguna columna de texto que no sea de meses.
    """
    import re
    resto = re.sub(r'^\d{1,2}\s*[-–.:)]*\s*', '', cod_str).strip()
    if len(resto) >= 3 and not resto.replace('.', '').replace(',', '').isdigit():
        return resto[:120]
    for j, val in enumerate(row):
        if j == codigo_col or j in mes_cols or val is None:
            continue
        s = str(val).strip()
        if len(s) >= 3 and not s.replace('.', '').replace(',', '').isdigit():
            return s[:120]
    return ''


def importar_excel_inei(filepath: str, area: str = '01',
                        anio_override: int | None = None,
                        serie: str | None = None) -> dict:
    """Importa valores de índices desde un Excel.

    Dos formatos, en orden:

    1. **El archivo oficial del INEI** — una hoja por mes y las columnas son
       ÁREAS GEOGRÁFICAS. Se leen todas las hojas y todas las áreas de una sola
       vez, y la serie sale de la base declarada en la cabecera.
    2. **Una planilla libre** — una hoja y las columnas son MESES. Es lo que
       arma un usuario a mano; ahí sí hace falta indicar el `area`.

    Distinguirlos importa: el lector anterior aplicaba SIEMPRE el criterio 2, y
    con el archivo oficial eso significa leer las áreas como si fueran meses.
    """
    try:
        import openpyxl
    except ImportError:
        return {'ok': False, 'msg': "openpyxl no instalado (pip install openpyxl)",
                'rows': [], 'ignorados': 0}

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return {'ok': False, 'msg': f"No se pudo abrir: {e}",
                'rows': [], 'ignorados': 0}

    oficial = _importar_oficial(wb, serie)
    if oficial:
        return oficial

    return _importar_libre(wb, area, anio_override, serie)


def _importar_libre(wb, area: str, anio_override: int | None,
                    serie: str | None) -> dict:
    """Planilla de formato libre: una sola hoja y las columnas son meses."""
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return {'ok': False, 'msg': "El archivo está vacío.",
                'rows': [], 'ignorados': 0}

    # Detectar año: buscar un texto tipo "ENERO 2026" o "2026"
    anio_detectado = anio_override
    if not anio_detectado:
        import re
        for row in data[:20]:
            for cell in row:
                if cell:
                    m = re.search(r'\b(20\d{2})\b', str(cell))
                    if m:
                        anio_detectado = int(m.group(1))
                        break
            if anio_detectado:
                break

    # Buscar fila de encabezado: la primera con al menos 6 valores que sean
    # meses parseables (o números 1..12)
    header_row_idx = -1
    mes_cols: dict[int, int] = {}   # idx_col → mes (1..12)
    codigo_col = -1
    for i, row in enumerate(data[:30]):
        mes_temp: dict[int, int] = {}
        codigo_temp = -1
        for j, val in enumerate(row):
            mes = _parse_mes(val)
            if mes:
                mes_temp[j] = mes
            elif val and 'codigo' in str(val).lower():
                codigo_temp = j
            elif val and ('indice' in str(val).lower() or 'iu' in str(val).lower()):
                if codigo_temp < 0:
                    codigo_temp = j
        if len(mes_temp) >= 6:
            header_row_idx = i
            mes_cols = mes_temp
            codigo_col = codigo_temp if codigo_temp >= 0 else 0
            break

    if header_row_idx < 0 or not mes_cols:
        return {
            'ok': False,
            'msg': ("No se encontró una tabla con encabezados de mes "
                    "(Ene/Feb/Mar… o 1/2/3…)."),
            'rows': [], 'ignorados': 0,
        }

    rows_out: list[dict] = []
    ignorados = 0
    codigos_encontrados: set[str] = set()
    nombres_encontrados: dict[str, str] = {}
    for row in data[header_row_idx + 1:]:
        if not row:
            continue
        # Código en col codigo_col, soportar valores tipo '01', 1, '1.0', '01 - Aceite'
        cod_raw = row[codigo_col] if codigo_col < len(row) else None
        if cod_raw is None:
            continue
        cod_str = str(cod_raw).strip()
        # Extraer dígitos iniciales
        import re
        m = re.match(r'^(\d{1,2})', cod_str)
        if not m:
            continue
        codigo = m.group(1).zfill(2)
        if not (1 <= int(codigo) <= 99):
            continue

        # Nombre del índice, para poder dar de alta los que el catálogo no
        # tenga con su descripción real y no con un «Índice 85» pelado.
        nombre = _nombre_en_fila(row, cod_str, codigo_col, mes_cols)
        if nombre:
            nombres_encontrados[codigo] = nombre

        for col_idx, mes in mes_cols.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == '':
                continue
            try:
                f = float(val)
                if f <= 0:
                    ignorados += 1
                    continue
            except Exception:
                ignorados += 1
                continue
            if not anio_detectado:
                ignorados += 1
                continue
            rows_out.append({
                'codigo': codigo, 'anio': anio_detectado,
                'mes': mes, 'area': area, 'valor': f,
                'serie': serie or serie_de(anio_detectado, mes),
                'nombre': nombres_encontrados.get(codigo, ''),
            })
            codigos_encontrados.add(codigo)

    return {
        'ok': True,
        'msg': f"OK — {len(rows_out)} valores listos para importar.",
        'rows': rows_out,
        'ignorados': ignorados,
        'anio_detectado': anio_detectado,
        'codigos_encontrados': codigos_encontrados,
        'nombres_encontrados': nombres_encontrados,
    }


# ─── Resoluciones mensuales del INEI (gob.pe) ────────────────────────────────
# El Excel de la base nueva es UN archivo que el INEI actualiza cuando quiere
# —a agosto de 2026 seguía con datos hasta marzo—, pero la resolución jefatural
# de cada mes SÍ sale puntual y se publica en gob.pe como PDF. De ahí salen los
# meses que al Excel le faltan.
GOBPE_IUPC = ("https://www.gob.pe/institucion/inei/informes-publicaciones/"
              "4025211-indices-unificados-de-precios-de-la-construccion-"
              "para-las-trece-areas-geograficas")

_MESES_NOMBRE = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
    'julio': 7, 'agosto': 8, 'septiembre': 9, 'setiembre': 9, 'octubre': 10,
    'noviembre': 11, 'diciembre': 12,
}


def _mes_del_titulo(titulo: str) -> tuple[int, int] | None:
    """(año, mes) del título de una resolución, o None si no lo dice.

    El título es lo único que identifica el período: «R.J. N°187-2026-INEI
    (ÍNDICES MES DE JULIO 2026)». El número de la resolución no sirve —el 016
    de 2026 es el cambio de base, no un mes.
    """
    import re
    t = (titulo or '').lower()
    mes = next((n for nombre, n in _MESES_NOMBRE.items() if nombre in t), None)
    anio = re.search(r'\b(20\d{2})\b', t)
    return (int(anio.group(1)), mes) if mes and anio else None


def buscar_resoluciones_gobpe(timeout: int = 20) -> list[dict]:
    """Resoluciones mensuales de índices publicadas en gob.pe.

    La página es HTML servido —no como el buscador de El Peruano, que es una
    aplicación de cliente y no se puede leer sin navegador— y enlaza los PDF
    con su mes en el título: «R.J. N°187-2026-INEI (ÍNDICES MES DE JULIO
    2026)». Publica el mes vigente y lo va reemplazando.

    Devuelve [{'url', 'resolucion', 'mes', 'anio', 'titulo'}].
    """
    import re
    import urllib.request

    req = urllib.request.Request(GOBPE_IUPC, headers={
        'User-Agent': ('Mozilla/5.0 (X11; Linux x86_64; rv:128.0) '
                       'Gecko/20100101 Firefox/128.0'),
        'Accept-Language': 'es-PE,es;q=0.9',
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        html = resp.read().decode('utf-8', errors='ignore')

    out: list[dict] = []
    vistos: set[str] = set()
    patron = re.compile(
        r'href="(https://cdn\.www\.gob\.pe/uploads/document/file/[^"]+?\.pdf)'
        r'[^"]*"[^>]*>.{0,400}?R\.J\.\s*N°?\s*([\d]+-\d{4}-INEI)\s*'
        r'\(([^)]{0,80})\)',
        re.S | re.I)
    for m in patron.finditer(html):
        url, resolucion, titulo = m.group(1), m.group(2), m.group(3)
        if url in vistos:
            continue
        t = titulo.lower()
        if 'indice' not in t and 'índice' not in t:
            continue          # mano de obra, factores de liquidación…
        periodo = _mes_del_titulo(titulo)
        if not periodo:
            continue
        vistos.add(url)
        out.append({'url': url, 'resolucion': resolucion, 'mes': periodo[1],
                    'anio': periodo[0],
                    'titulo': ' '.join(titulo.split())})
    out.sort(key=lambda r: (r['anio'], r['mes']), reverse=True)
    return out


def importar_pdf_resolucion(filepath: str, serie: str | None = None) -> dict:
    """Lee la tabla de índices de una resolución jefatural en PDF.

    El formato es fijo: una fila por código y una columna por área geográfica
    —13 en la base 2025, 6 en la de 1992—, con «(*)» donde el índice no existe
    en esa área y coma decimal.

    Del propio texto salen el mes y la base, así que no hace falta que el
    usuario los indique.
    """
    import re
    try:
        import pdfplumber
    except ImportError:
        return {'ok': False, 'msg': "pdfplumber no instalado.",
                'rows': [], 'ignorados': 0}
    try:
        with pdfplumber.open(filepath) as pdf:
            txt = "\n".join((p.extract_text() or '') for p in pdf.pages)
    except Exception as e:
        return {'ok': False, 'msg': f"No se pudo leer el PDF: {e}",
                'rows': [], 'ignorados': 0}

    bajo = txt.lower()
    if serie is None:
        serie = SERIE_2025 if 'diciembre 2025' in bajo else (
            SERIE_1992 if 'julio 1992' in bajo else SERIE_ACTUAL)

    mes = anio = None
    m = re.search(r'mes de\s+([a-záéíóú]+)\s+(?:de\s+)?(20\d{2})', bajo)
    if not m:
        m = re.search(r'\b(' + '|'.join(_MESES_NOMBRE) + r')\s+(?:de\s+)?(20\d{2})',
                      bajo)
    if m:
        mes = _MESES_NOMBRE.get(m.group(1))
        anio = int(m.group(2))
    if not mes or not anio:
        return {'ok': False, 'rows': [], 'ignorados': 0,
                'msg': "No pude identificar el mes de la resolución en el PDF."}

    rows: list[dict] = []
    ignorados = 0
    codigos: set[str] = set()
    fila_re = re.compile(r'^\s*(\d{1,2})\s+((?:(?:[\d.]+,\d+|\(\*\))\s*)+)$')
    for linea in txt.splitlines():
        f = fila_re.match(linea)
        if not f:
            continue
        codigo = f.group(1).zfill(2)
        celdas = f.group(2).split()
        if len(celdas) < 4:      # una fila de índices trae 6 o 13 columnas
            continue
        for i, celda in enumerate(celdas, start=1):
            if celda == '(*)':
                continue         # el índice no existe en esa área: no es cero
            try:
                valor = float(celda.replace('.', '').replace(',', '.'))
            except ValueError:
                ignorados += 1
                continue
            if valor <= 0:
                ignorados += 1
                continue
            rows.append({'codigo': codigo, 'serie': serie, 'anio': anio,
                         'mes': mes, 'area': f"{i:02d}", 'valor': valor})
            codigos.add(codigo)

    if not rows:
        return {'ok': False, 'rows': [], 'ignorados': ignorados,
                'msg': "El PDF no trae una tabla de índices reconocible."}

    areas = sorted({r['area'] for r in rows})
    return {
        'ok': True,
        'formato': 'resolucion_pdf',
        'serie': serie,
        'msg': (f"Resolución del INEI: {len(rows)} valores · "
                f"{len(codigos)} índices · {len(areas)} áreas · "
                f"{anio}-{mes:02d}."),
        'rows': rows,
        'ignorados': ignorados,
        'anio_detectado': anio,
        'mes_detectado': mes,
        'codigos_encontrados': codigos,
        'nombres_encontrados': {},
        'areas': areas,
        'periodos': 1,
    }


ELPERUANO_DISPOSITIVO = "https://busquedas.elperuano.pe/dispositivo/NL/{id}-1"


URL_INDICES_PUBLICADOS = (
    "https://raw.githubusercontent.com/ingelibre/ingepresupuestos/main/"
    "app/resources/indices_inei_valores.json.gz")


def _filas_de_bloque(serie: str, bloque: dict) -> list[dict]:
    """Del formato compacto (un arreglo por área) a filas sueltas."""
    areas = bloque.get('areas') or []
    filas: list[dict] = []
    for periodo, codigos in (bloque.get('datos') or {}).items():
        try:
            anio, mes = (int(x) for x in periodo.split('-'))
        except ValueError:
            continue
        for codigo, valores in codigos.items():
            for i, valor in enumerate(valores):
                if valor is None or i >= len(areas):
                    continue
                filas.append({'codigo': codigo, 'serie': serie, 'anio': anio,
                              'mes': mes, 'area': areas[i],
                              'valor': float(valor)})
    return filas


def descargar_indices_publicados(timeout: int = 30) -> dict:
    """Trae el histórico que la Action del repositorio mantiene al día.

    Es la fuente preferida y la más completa: un solo archivo con las dos
    bases, ya reconciliado desde el Excel del INEI, los PDF de gob.pe y las
    resoluciones de El Peruano. Se actualiza dos veces al mes por su cuenta, así
    que la app no depende de que salga una versión nueva para tener el mes
    pasado.

    Las otras fuentes siguen ahí a propósito: si el repositorio no responde,
    o si el INEI publica antes de que la Action corra, sincronizar igual trae
    lo que haya.
    """
    import gzip
    import io
    import json as _json
    import urllib.request

    req = urllib.request.Request(URL_INDICES_PUBLICADOS, headers={
        'User-Agent': 'IngePresupuestos/indices',
        'Accept-Encoding': 'identity',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            crudo = resp.read(20 * 1024 * 1024)
    except Exception as e:
        return {'ok': False, 'rows': [], 'ignorados': 0,
                'url': URL_INDICES_PUBLICADOS,
                'msg': f"No se pudo leer el histórico publicado: {e}"}
    try:
        with gzip.open(io.BytesIO(crudo), 'rt', encoding='utf-8') as fh:
            doc = _json.load(fh)
    except Exception as e:
        return {'ok': False, 'rows': [], 'ignorados': 0,
                'url': URL_INDICES_PUBLICADOS,
                'msg': f"El histórico publicado no se pudo leer: {e}"}

    filas: list[dict] = []
    for serie, bloque in (doc.get('series') or {}).items():
        filas += _filas_de_bloque(serie, bloque)
    if not filas:
        return {'ok': False, 'rows': [], 'ignorados': 0,
                'url': URL_INDICES_PUBLICADOS,
                'msg': "El histórico publicado vino vacío."}
    return {
        'ok': True, 'rows': filas, 'ignorados': 0,
        'url': URL_INDICES_PUBLICADOS,
        'generado': doc.get('generado', ''),
        'tamano_kb': round(len(crudo) / 1024, 1),
        'codigos_encontrados': {f['codigo'] for f in filas},
        'msg': (f"Histórico publicado ({doc.get('generado', 'sin fecha')}): "
                f"{len(filas)} valores."),
    }


def importar_html_elperuano(url: str, serie: str | None = None,
                            timeout: int = 45) -> dict:
    """Lee la tabla de índices de una resolución publicada en El Peruano.

    Es la tercera fuente, y la que cubre lo que a las otras dos se les escapa:
    el INEI congela su Excel durante meses y gob.pe solo deja el PDF del mes
    vigente, pero **El Peruano conserva todas las resoluciones publicadas**.

    Su BUSCADOR no sirve —`busquedas.elperuano.pe` y `/cuadernillo/NL/` son
    aplicaciones de cliente y sin navegador devuelven cero—, pero la página de
    cada dispositivo (`/dispositivo/NL/<id>-1`) viene servida con el texto
    íntegro y la tabla en HTML, códigos por fila y las áreas por columna, con
    «(*)» donde el índice no existe.

    Sirve además como verificación: ante una duda sobre un valor, esta es la
    publicación oficial, no una copia.
    """
    import re
    import urllib.request

    if not str(url).lower().startswith(('http://', 'https://')):
        return {'ok': False, 'msg': "URL no válida.", 'rows': [], 'ignorados': 0}
    req = urllib.request.Request(url, headers={
        'User-Agent': ('Mozilla/5.0 (X11; Linux x86_64; rv:128.0) '
                       'Gecko/20100101 Firefox/128.0'),
        'Accept-Language': 'es-PE,es;q=0.9',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            html = resp.read().decode('utf-8', errors='replace')
    except Exception as e:
        return {'ok': False, 'msg': f"No se pudo abrir El Peruano: {e}",
                'rows': [], 'ignorados': 0, 'url': url}

    texto = re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', ' ', html))
    m = re.search(r'mes de\s+([a-zA-ZáéíóúÁÉÍÓÚ]+)\s+de\s+(20\d{2})', texto)
    periodo = _mes_del_titulo(m.group(0)) if m else None
    if not periodo:
        return {'ok': False, 'rows': [], 'ignorados': 0, 'url': url,
                'msg': "No pude identificar el mes de la resolución."}
    anio, mes = periodo
    if serie is None:
        bajo = texto.lower()
        serie = (SERIE_2025 if 'diciembre 2025' in bajo else
                 SERIE_1992 if 'julio 1992' in bajo else serie_de(anio, mes))

    tablas = re.findall(r'<table.*?</table>', html, re.S | re.I)
    if not tablas:
        return {'ok': False, 'rows': [], 'ignorados': 0, 'url': url,
                'msg': "La página no trae la tabla de índices."}

    def celdas(fila):
        return [re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', '', c)).strip()
                for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', fila, re.S | re.I)]

    rows: list[dict] = []
    ignorados = 0
    codigos: set[str] = set()
    for fila in re.findall(r'<tr.*?</tr>', max(tablas, key=len), re.S | re.I):
        cs = celdas(fila)
        if len(cs) < 2 or not cs[0].strip().isdigit():
            continue                      # cabecera «Cód. 1 2 3…» y ruido
        codigo = cs[0].strip().zfill(2)
        for i, bruto in enumerate(cs[1:], start=1):
            v = bruto.strip()
            if not v or '(*)' in v:
                continue                  # el índice no existe en esa área
            try:
                valor = float(v.replace('.', '').replace(',', '.'))
            except ValueError:
                ignorados += 1
                continue
            if valor <= 0:
                ignorados += 1
                continue
            rows.append({'codigo': codigo, 'serie': serie, 'anio': anio,
                         'mes': mes, 'area': f"{i:02d}", 'valor': valor})
            codigos.add(codigo)

    if not rows:
        return {'ok': False, 'rows': [], 'ignorados': ignorados, 'url': url,
                'msg': "No se reconoció ningún valor en la tabla."}
    areas = {r['area'] for r in rows}
    return {
        'ok': True, 'rows': rows, 'ignorados': ignorados, 'url': url,
        'serie': serie, 'codigos_encontrados': codigos,
        'msg': (f"El Peruano: {len(rows)} valores · {len(codigos)} índices · "
                f"{len(areas)} áreas · {anio}-{mes:02d}."),
    }


def descargar_resolucion_gobpe(url: str, serie: str | None = None,
                               timeout: int = 60) -> dict:
    """Descarga una resolución en PDF desde gob.pe y lee su tabla."""
    import tempfile
    import urllib.request

    if not url.lower().startswith(('http://', 'https://')):
        return {'ok': False, 'msg': "URL no válida.", 'rows': [],
                'ignorados': 0}
    req = urllib.request.Request(url, headers={
        'User-Agent': ('Mozilla/5.0 (X11; Linux x86_64; rv:128.0) '
                       'Gecko/20100101 Firefox/128.0'),
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            data = resp.read(25 * 1024 * 1024)
    except Exception as e:
        return {'ok': False, 'msg': f"No se pudo descargar: {e}",
                'rows': [], 'ignorados': 0, 'url': url}

    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        tmp.write(data)
        ruta = tmp.name
    try:
        res = importar_pdf_resolucion(ruta, serie=serie)
    finally:
        import os as _os
        try:
            _os.unlink(ruta)
        except OSError:
            pass
    res['url'] = url
    res['tamano_kb'] = round(len(data) / 1024, 1)
    return res


def exportar_json(filepath: str, area: str | None = None) -> int:
    """Exporta toda la serie a JSON. Si ``area`` no es None, filtra por ella.
    Retorna el número de valores exportados."""
    conn = get_db()
    try:
        q = ("SELECT codigo, anio, mes, area, valor "
             "FROM indices_inei_valores")
        p = []
        if area:
            q += " WHERE area=?"; p.append(area)
        q += " ORDER BY codigo, anio, mes, area"
        rows = [dict(r) for r in conn.execute(q, p).fetchall()]
    finally:
        conn.close()
    payload = {
        'version': 1,
        'tipo': 'indices_inei',
        'area_filtro': area,
        'valores': rows,
    }
    Path(filepath).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding='utf-8'
    )
    return len(rows)


def importar_json(filepath: str) -> dict:
    """Importa valores desde JSON exportado por ``exportar_json``."""
    try:
        payload = json.loads(Path(filepath).read_text(encoding='utf-8'))
    except Exception as e:
        return {'ok': False, 'msg': f"Archivo inválido: {e}", 'n_ok': 0, 'n_err': 0}
    if payload.get('tipo') != 'indices_inei':
        return {'ok': False, 'msg': "El archivo no es de tipo indices_inei.",
                'n_ok': 0, 'n_err': 0}
    rows = payload.get('valores') or []
    ok, err = guardar_valores(rows)
    return {'ok': True, 'msg': f"{ok} valores importados, {err} ignorados",
            'n_ok': ok, 'n_err': err}


# ─── Descarga por URL ────────────────────────────────────────────────────────
def descargar_desde_url(url: str, area: str = '01',
                        anio_override: int | None = None,
                        serie: str | None = None) -> dict:
    """Descarga desde una URL pública y la parsea según lo que sea.

    Tres formas, porque el usuario puede llegar con cualquiera de las tres
    fuentes oficiales: el **Excel** del INEI, el **PDF** de una resolución en
    gob.pe, o la página de la resolución en **El Peruano** —que es la que
    conserva los meses viejos y sirve para salir de dudas sobre un valor—.

    Acepta solo http/https. Timeout de 30s. Tamaño máximo: 20 MB.
    Retorna el mismo dict que ``importar_excel_inei`` más:
        - 'url': URL fuente
        - 'tamano_kb': tamaño del archivo descargado
    """
    import tempfile

    bajo = str(url).lower()
    if 'elperuano.pe' in bajo:
        return importar_html_elperuano(url, serie=serie)
    if bajo.split('?')[0].endswith('.pdf'):
        return descargar_resolucion_gobpe(url, serie=serie)
    import urllib.request
    import urllib.error

    if not url:
        return {'ok': False, 'msg': "URL vacía.", 'rows': [], 'ignorados': 0}
    if not (url.startswith("http://") or url.startswith("https://")):
        return {'ok': False, 'msg': "La URL debe empezar con http:// o https://",
                'rows': [], 'ignorados': 0}

    # Headers que muchos servidores requieren para no bloquearnos
    req = urllib.request.Request(url, headers={
        'User-Agent': ('Mozilla/5.0 (X11; Linux x86_64) ingePresupuestos/1.0'),
        'Accept': '*/*',
    })

    tmp_path: str | None = None
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            ct = resp.headers.get('Content-Type', '').lower()
            cl = resp.headers.get('Content-Length')
            if cl and int(cl) > 20 * 1024 * 1024:
                return {'ok': False,
                        'msg': f"Archivo demasiado grande ({int(cl)//1024} KB).",
                        'rows': [], 'ignorados': 0}

            # Si la URL apunta a HTML, lo más probable es que sea una página y
            # no el Excel directo
            if 'html' in ct and not url.lower().endswith(('.xlsx', '.xls')):
                return {
                    'ok': False,
                    'msg': ("La URL devuelve HTML, no un Excel. Abre la "
                            "página en el navegador, haz clic derecho sobre "
                            "el enlace del Excel y elige «Copiar dirección "
                            "del enlace», luego pégala aquí."),
                    'rows': [], 'ignorados': 0,
                }

            data = resp.read()

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.write(data)
        tmp.close()
        tmp_path = tmp.name

        res = importar_excel_inei(tmp_path, area=area,
                                  anio_override=anio_override, serie=serie)
        res['url'] = url
        res['tamano_kb'] = round(len(data) / 1024, 1)
        return res

    except urllib.error.HTTPError as e:
        return {'ok': False, 'msg': f"Error HTTP {e.code} al descargar.",
                'rows': [], 'ignorados': 0}
    except urllib.error.URLError as e:
        return {'ok': False, 'msg': f"Error de conexión: {e.reason}",
                'rows': [], 'ignorados': 0}
    except Exception as e:
        return {'ok': False, 'msg': f"Error inesperado: {e}",
                'rows': [], 'ignorados': 0}
    finally:
        if tmp_path:
            try:
                import os
                os.unlink(tmp_path)
            except Exception:
                pass


# ─── Pegar desde portapapeles ────────────────────────────────────────────────
def importar_desde_texto(texto: str, area: str = '01',
                          anio_override: int | None = None) -> dict:
    """Parsea contenido tabular (CSV, TSV o tabla pegada desde Excel/web/PDF).

    Soporta los mismos formatos que ``importar_excel_inei`` pero recibe el
    contenido como texto crudo. Detecta separador automáticamente
    (tab, coma, punto y coma, múltiples espacios).
    """
    import csv
    import io
    import re
    import tempfile

    if not texto or not texto.strip():
        return {'ok': False, 'msg': "Nada para pegar.",
                'rows': [], 'ignorados': 0}

    # Detectar separador por mayoría en las primeras 5 líneas
    primeras = "\n".join(texto.splitlines()[:5])
    counts = {
        '\t': primeras.count('\t'),
        ';':  primeras.count(';'),
        ',':  primeras.count(','),
    }
    sep = max(counts, key=counts.get)
    if counts[sep] == 0:
        # Sin separador estándar, tratar de detectar espacios múltiples
        sep_re = re.compile(r' {2,}|\t')
        rows = [sep_re.split(line.strip()) for line in texto.splitlines()
                if line.strip()]
    else:
        rows = list(csv.reader(io.StringIO(texto), delimiter=sep))

    if not rows:
        return {'ok': False, 'msg': "No se pudieron parsear filas.",
                'rows': [], 'ignorados': 0}

    # Escribimos un Excel temporal en memoria para reutilizar el parser robusto
    try:
        import openpyxl
    except ImportError:
        return {'ok': False, 'msg': "openpyxl no instalado.",
                'rows': [], 'ignorados': 0}

    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append([c.strip() if isinstance(c, str) else c for c in r])

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    try:
        wb.save(tmp.name)
        res = importar_excel_inei(tmp.name, area=area,
                                  anio_override=anio_override)
        res['filas_pegadas'] = len(rows)
        return res
    finally:
        try:
            import os
            os.unlink(tmp.name)
        except Exception:
            pass


# ─── Auto-detección del último archivo INEI ──────────────────────────────────
INEI_BASE = "https://www.inei.gob.pe/media/MenuRecursivo/indices_tematicos"
INEI_PATTERN = "06_indices_unificados_de_precios_de_la_construccion_{mes}{ano}.xlsx"
INEI_PATTERN_NUEVA = "07_indices_unificados_de_precios_de_la_construccion_1.xlsx"

_MESES_CORTOS_INEI = ['ene', 'feb', 'mar', 'abr', 'may', 'jun',
                      'jul', 'ago', 'set', 'oct', 'nov', 'dic']


def buscar_ultimo_excel_inei(serie: str = SERIE_ACTUAL) -> dict:
    """Busca por HEAD el último Excel publicado por el INEI para una serie.

    Son DOS archivos distintos y hasta ahora solo se probaba el viejo:

    * base **Diciembre 2025** — `07_..._1.xlsx`, un único archivo con las 13
      áreas que el INEI va ampliando mes a mes;
    * base **Julio 1992** — `06_..._{mes}{año}.xlsx`, uno por mes y acumulativo
      (el de dic-25 trae desde 2013). Se prueba hacia atrás desde el mes
      actual hasta dar con uno que responda 200.

    Sincronizar traía siempre el segundo, así que los valores entraban en la
    serie 1992 y en la vista —que muestra la vigente— no aparecía nada.

    Retorna::

        {'ok': bool, 'url': str|None, 'msg': str, 'mes_detectado': str|None,
         'anio_detectado': int|None, 'serie': str}
    """
    import urllib.request
    from datetime import date

    headers = {
        'User-Agent': ('Mozilla/5.0 (X11; Linux x86_64) '
                       'ingePresupuestos/1.0'),
    }

    if serie == SERIE_2025:
        url = f"{INEI_BASE}/{INEI_PATTERN_NUEVA}"
        try:
            req = urllib.request.Request(url, method='HEAD', headers=headers)
            with urllib.request.urlopen(req, timeout=8) as resp:
                if resp.status == 200:
                    return {'ok': True, 'url': url, 'serie': SERIE_2025,
                            'msg': "Archivo del INEI con la base "
                                   "Diciembre 2025 = 100 (13 áreas).",
                            'mes_detectado': None, 'anio_detectado': None}
        except Exception as e:
            return {'ok': False, 'url': None, 'serie': SERIE_2025,
                    'msg': f"No se pudo alcanzar el archivo del INEI: {e}",
                    'mes_detectado': None, 'anio_detectado': None}
        return {'ok': False, 'url': None, 'serie': SERIE_2025,
                'msg': "El INEI no responde con el archivo de la base 2025.",
                'mes_detectado': None, 'anio_detectado': None}

    hoy = date.today()
    # Probar desde el mes actual hasta 18 meses atrás
    candidatos: list[tuple[int, int]] = []
    y, m = hoy.year, hoy.month
    for _ in range(18):
        candidatos.append((y, m))
        m -= 1
        if m < 1:
            m = 12
            y -= 1

    for y, m in candidatos:
        mes_str = _MESES_CORTOS_INEI[m - 1]
        ano_str = str(y % 100).zfill(2)
        fname = INEI_PATTERN.format(mes=mes_str, ano=ano_str)
        url = f"{INEI_BASE}/{fname}"
        try:
            req = urllib.request.Request(url, method='HEAD', headers=headers)
            with urllib.request.urlopen(req, timeout=6) as resp:
                if resp.status == 200:
                    return {
                        'ok': True,
                        'url': url,
                        'serie': SERIE_1992,
                        'msg': (f"Archivo del INEI de {mes_str.title()} {y} "
                                f"(base Julio 1992 = 100, acumulativo)."),
                        'mes_detectado': mes_str,
                        'anio_detectado': y,
                    }
        except Exception:
            continue

    return {
        'ok': False,
        'url': None,
        'serie': SERIE_1992,
        'msg': ("No se encontró ningún archivo INEI con el patrón conocido en "
                "los últimos 18 meses. Tu conexión puede estar bloqueada o el "
                "INEI cambió el formato de URL."),
        'mes_detectado': None,
        'anio_detectado': None,
    }


def descargar_ultimo_inei(area: str = '01',
                          serie: str = SERIE_ACTUAL) -> dict:
    """Busca el Excel del INEI de esa serie y lo importa.

    `area` solo se usa si el archivo resultara ser una planilla de formato
    libre: el archivo oficial trae TODAS las áreas y se leen todas.
    """
    busq = buscar_ultimo_excel_inei(serie)
    if not busq['ok']:
        return {'ok': False, 'msg': busq['msg'], 'rows': [], 'ignorados': 0,
                'url': None, 'serie': serie}
    res = descargar_desde_url(busq['url'], area=area,
                              anio_override=busq['anio_detectado'],
                              serie=busq.get('serie', serie))
    res['mes_detectado'] = busq['mes_detectado']
    res['anio_detectado_url'] = busq['anio_detectado']
    res.setdefault('serie', busq.get('serie', serie))
    return res
